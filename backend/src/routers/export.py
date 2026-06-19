from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Any
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.dependencies.auth import get_current_user

router = APIRouter()


class ExportRequest(BaseModel):
    """Request body for Excel export."""
    columns: List[str]
    rows: List[List[Any]]
    filename: str = "result"


@router.post("/excel")
async def export_to_excel(request: ExportRequest, user=Depends(get_current_user)):
    """
    Accepts columns and rows from a query result and returns
    a formatted .xlsx file as a downloadable stream.
    """
    if not request.columns:
        raise HTTPException(status_code=400, detail="Columns list cannot be empty")

    wb = Workbook()
    ws = wb.active
    ws.title = "Query Results"

    # --- Header styling ---
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    ws.append(request.columns)
    for col_idx, _ in enumerate(request.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Data rows ---
    data_alignment = Alignment(vertical="center")
    for row_data in request.rows:
        ws.append(row_data)
        current_row = ws.max_row
        for col_idx in range(1, len(request.columns) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = thin_border
            cell.alignment = data_alignment

    # Auto-fit column widths
    for col_idx, column_name in enumerate(request.columns, start=1):
        max_length = len(str(column_name))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Write to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_filename = "".join(c for c in request.filename if c.isalnum() or c in ("-", "_")) or "result"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{safe_filename}.xlsx"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
