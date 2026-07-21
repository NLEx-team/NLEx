import logging
import os
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart, Reference, Series as ChartSeries
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
import xlsxwriter

logger = logging.getLogger(__name__)

EXPORTS_DIR = os.environ.get("EXPORTS_DIR", "/app/exports")

# How many data rows to sample for auto-width calculation (keeps it fast).
_WIDTH_SAMPLE_ROWS = 200
_MAX_COL_WIDTH = 60
_MIN_COL_WIDTH = 10


class ExcelExportService:
    """
    Service responsible for generating and caching Excel files
    from chat query results.

    Uses xlsxwriter for high-performance streaming writes — 3-5x faster
    than openpyxl and avoids the costly post-processing step.
    """

    def __init__(self, exports_dir: str = EXPORTS_DIR):
        self.exports_dir = exports_dir
        os.makedirs(self.exports_dir, exist_ok=True)

    def _get_file_path(self, export_id: str) -> str:
        """Returns the file path for a given export_id."""
        return os.path.join(self.exports_dir, f"{export_id}.xlsx")

    def get_cached_file(self, export_id: str) -> Optional[str]:
        """
        Returns the path to a cached Excel file if it exists,
        otherwise None.
        """
        path = self._get_file_path(export_id)
        if os.path.exists(path):
            return path
        return None

    def _get_metadata_path(self, export_id: str) -> str:
        return os.path.join(self.exports_dir, f"{export_id}.meta.json")

    def save_export_metadata(
        self,
        export_id: str,
        sql: str,
        headers: List[str],
        catalog_mapping: Dict[str, str] = None,
        filename: str = None,
        chart: Optional[Dict[str, any]] = None
    ) -> None:
        import json
        meta_path = self._get_metadata_path(export_id)
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump({
                "sql": sql, 
                "headers": headers,
                "catalog_mapping": catalog_mapping or {},
                "filename": filename or f"export_{export_id}",
                "chart": chart
            }, f)

    def _find_column_index(self, headers: List[str], column_name: str) -> Optional[int]:
        """Find 1-based column index by header name."""
        for i, h in enumerate(headers):
            if h == column_name:
                return i + 1
        return None

    def _add_chart_to_workbook(self, wb: Workbook, ws: any, chart_spec: dict, headers: List[str], row_count: int):
        """Add a chart to the workbook based on chart spec from LLM."""
        chart_type = chart_spec.get("type", "bar")
        title = chart_spec.get("title", "")

        last_data_row = row_count + 1  # row 1 is header

        # Map column names to 1-based indices
        if chart_type == "pie":
            cat_col = self._find_column_index(headers, chart_spec.get("category_column", ""))
            val_col = self._find_column_index(headers, chart_spec.get("value_column", ""))
            if not cat_col or not val_col:
                logger.warning(f"Cannot create pie chart: columns not found")
                return
            chart = PieChart()
            chart.title = title
            data_ref = Reference(ws, min_col=val_col, min_row=1, max_row=last_data_row)
            cats_ref = Reference(ws, min_col=cat_col, min_row=2, max_row=last_data_row)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.dataLabels = DataLabelList()
            chart.dataLabels.showPercent = True
            chart.dataLabels.showCatName = True

        elif chart_type == "scatter":
            x_col = self._find_column_index(headers, chart_spec.get("x_column", ""))
            y_cols = [self._find_column_index(headers, y) for y in chart_spec.get("y_columns", [])]
            y_cols = [c for c in y_cols if c]
            if not x_col or not y_cols:
                logger.warning("Cannot create scatter chart: columns not found")
                return
            chart = ScatterChart()
            chart.title = title
            chart.x_axis.title = chart_spec.get("x_column", "")
            chart.y_axis.title = chart_spec.get("y_columns", [""])[0] if chart_spec.get("y_columns") else ""
            for y_col in y_cols:
                x_vals = Reference(ws, min_col=x_col, min_row=2, max_row=last_data_row)
                y_vals = Reference(ws, min_col=y_col, min_row=1, max_row=last_data_row)
                series = ChartSeries(y_vals, x_vals, title_from_data=True)
                chart.series.append(series)

        else:
            # bar, line, area
            x_col = self._find_column_index(headers, chart_spec.get("x_column", ""))
            y_cols = [self._find_column_index(headers, y) for y in chart_spec.get("y_columns", [])]
            y_cols = [c for c in y_cols if c]
            if not x_col or not y_cols:
                logger.warning(f"Cannot create {chart_type} chart: columns not found")
                return

            if chart_type == "bar":
                chart = BarChart()
                chart.type = "col"
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "area":
                chart = AreaChart()
            else:
                chart = BarChart()
                chart.type = "col"

            chart.title = title
            chart.x_axis.title = chart_spec.get("x_column", "")
            chart.y_axis.title = chart_spec.get("y_columns", [""])[0] if chart_spec.get("y_columns") else ""

            if chart_spec.get("stacked"):
                chart.grouping = "stacked"

            for y_col in y_cols:
                data_ref = Reference(ws, min_col=y_col, min_row=1, max_row=last_data_row)
                chart.add_data(data_ref, titles_from_data=True)

            cats_ref = Reference(ws, min_col=x_col, min_row=2, max_row=last_data_row)
            chart.set_categories(cats_ref)

        chart.style = 10
        chart.width = 25
        chart.height = 15

        ws_chart = wb.create_sheet(title="Chart")
        ws_chart.add_chart(chart, "A1")

    def generate_and_get_excel(self, export_id: str, db_service) -> str:
        file_path = self._get_file_path(export_id)
        if os.path.exists(file_path):
            return file_path
            
        import json
        meta_path = self._get_metadata_path(export_id)
        if not os.path.exists(meta_path):
            raise ValueError("Export metadata not found.")
            
        with open(meta_path, "r", encoding="utf-8") as f:
            meta = json.load(f)
            
        sql = meta["sql"]
        headers = meta["headers"]
        catalog_mapping = meta.get("catalog_mapping", {})

        chart_spec = meta.get("chart")
        
        # Use WriteOnlyWorkbook for memory efficiency
        wb = Workbook(write_only=True)
        # in develop it was:
        # wb = xlsxwriter.Workbook(file_path, {"constant_memory": True})
        
        ws = wb.create_sheet("Query Results")

        # --- Header styling ---
        from openpyxl.cell import WriteOnlyCell
        
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        data_alignment = Alignment(vertical="center")

        # Track max width per column (seeded with header lengths)
        col_widths = [min(len(str(h)) + 4, _MAX_COL_WIDTH) for h in headers]

        # Write header row
        header_cells = []
        for h in headers:
            c = WriteOnlyCell(ws, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_alignment
            c.border = thin_border
            header_cells.append(c)
        ws.append(header_cells)

        # Stream data from DB via the read-only guarded path: export must never
        # execute anything other than a single SELECT/CTE, even on re-run.
        row_idx = 1
        row_count = 0
        for row_data in db_service.execute_readonly_sync_stream(sql, chunk_size=2000):
            row_idx += 1
            data_cells = []
            for ci, val in enumerate(row_data):
                # Keep numeric types for chart support; stringify unsupported types
                if val is not None and not isinstance(val, (int, float, str, bool)):
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        val = str(val)
                
                if isinstance(val, str) and catalog_mapping:
                    for t_name, d_name in catalog_mapping.items():
                        if t_name in val:
                            val = val.replace(t_name, d_name)

                # Sample first N rows for auto-width
                if row_idx <= _WIDTH_SAMPLE_ROWS and ci < len(col_widths):
                    cell_len = len(str(val)) + 2 if val is not None else 0
                    if cell_len > col_widths[ci]:
                        col_widths[ci] = min(cell_len, _MAX_COL_WIDTH)

                cell = WriteOnlyCell(ws, value=val)
                cell.border = thin_border
                cell.alignment = data_alignment
                data_cells.append(cell)
            ws.append(data_cells)
            row_count += 1
            
        wb.save(file_path)

        # --- Post-process: apply auto-fitted widths + freeze header row + add chart ---
        try:
            wb2 = load_workbook(file_path)
            ws2 = wb2.active
            for ci, width in enumerate(col_widths):
                ws2.column_dimensions[get_column_letter(ci + 1)].width = max(width, _MIN_COL_WIDTH)
            ws2.freeze_panes = "A2"

            # Add chart if a chart spec is present
            if chart_spec and row_count > 0:
                try:
                    self._add_chart_to_workbook(wb2, ws2, chart_spec, headers, row_count)
                except Exception as chart_err:
                    logger.error(f"Chart generation failed: {chart_err}", exc_info=True)

            wb2.save(file_path)
        except Exception as e:
            logger.warning(f"Post-processing failed (file is still valid): {e}")

        row_count = row_idx - 1
        logger.info(f"Excel file streamed and saved: {file_path} ({row_count} rows)")
        return file_path
